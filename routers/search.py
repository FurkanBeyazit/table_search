import os
import io
import tempfile
from urllib.parse import quote
from fastapi import APIRouter, Query, HTTPException, Body
from fastapi.responses import FileResponse, StreamingResponse
from typing import List
from datetime import datetime, timedelta

import database
from config import (
    BHVR_TABLE, DST_TABLE, EVENT_COL,
    BHVR_EVENTS, DST_EVENTS, ALL_EVENTS,
    LINUX_PATH_PREFIX, WINDOWS_MOUNT_LETTER, API_BASE_URL,
    REPORT_API_URL,
)

router = APIRouter(prefix="/api/search", tags=["search"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_dt_param(s: str) -> str:
    """Accept both '20260316111320' (14-digit) and '2026-03-16 11:13:20' formats."""
    s = s.strip()
    if len(s) == 14 and s.isdigit():
        return datetime.strptime(s, "%Y%m%d%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
    return s


def linux_to_win(img_path: str) -> str:
    """Convert Linux path to Windows path string."""
    win = img_path.replace(LINUX_PATH_PREFIX, f"{WINDOWS_MOUNT_LETTER}:\\")
    return win.replace("/", "\\")


def img_to_api_url(img_path: str) -> str:
    if not img_path:
        return ""
    return f"{API_BASE_URL}/api/search/image?path={quote(img_path)}"


def img_to_thumb_url(img_path: str) -> str:
    if not img_path:
        return ""
    return f"{API_BASE_URL}/api/search/thumbnail?path={quote(img_path)}"


def parse_dtct_dt(val) -> str:
    if not val:
        return ""
    try:
        dt = datetime.strptime(str(val).strip(), "%Y%m%d%H%M%S")
        return (dt + timedelta(hours=9)).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return str(val)


def build_base_sql(events: List[str], start_dt: str, end_dt: str, node_ids: List[str] = None):
    bhvr = [e for e in events if e in BHVR_EVENTS]
    dst  = [e for e in events if e in DST_EVENTS]
    parts, params = [], []

    node_clause = ""
    node_params = []
    if node_ids:
        ph = ", ".join(["%s"] * len(node_ids))
        node_clause = f" AND CAST(node_id AS TEXT) IN ({ph})"
        node_params = node_ids

    if bhvr:
        ph = ", ".join(["%s"] * len(bhvr))
        parts.append(
            f"SELECT node_id, ch, reg_dt, dtct_dt, img_path, {EVENT_COL} AS event_type, NULL::float AS dst_val "
            f"FROM {BHVR_TABLE} "
            f"WHERE reg_dt BETWEEN %s AND %s AND {EVENT_COL} IN ({ph}){node_clause}"
        )
        params += [start_dt, end_dt] + bhvr + node_params

    if dst:
        ph = ", ".join(["%s"] * len(dst))
        parts.append(
            f"SELECT node_id, ch, reg_dt, dtct_dt, img_path, {EVENT_COL} AS event_type, dst_val "
            f"FROM {DST_TABLE} "
            f"WHERE reg_dt BETWEEN %s AND %s AND {EVENT_COL} IN ({ph}){node_clause}"
        )
        params += [start_dt, end_dt] + dst + node_params

    if not parts:
        return None, []

    return " UNION ALL ".join(parts), params


# ── Image endpoints ───────────────────────────────────────────────────────────

@router.get("/image")
def get_image(path: str):
    """Full-size image served over HTTP."""
    win_path = linux_to_win(path)
    if not os.path.isfile(win_path):
        raise HTTPException(status_code=404, detail=f"Not found: {win_path}")
    return FileResponse(win_path)


@router.get("/thumbnail")
def get_thumbnail(path: str):
    """Resized thumbnail (120×90, JPEG q50) with browser cache."""
    try:
        from PIL import Image
    except ImportError:
        raise HTTPException(status_code=500, detail="Pillow not installed")

    win_path = linux_to_win(path)
    if not os.path.isfile(win_path):
        raise HTTPException(status_code=404, detail=f"Not found: {win_path}")

    img = Image.open(win_path)
    img.thumbnail((120, 90), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=50)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="image/jpeg",
        headers={"Cache-Control": "public, max-age=3600"},
    )


# ── Search endpoints ──────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats(
    start_dt: str,
    end_dt: str,
    events: List[str] = Query(default=ALL_EVENTS),
    node_ids: List[str] = Query(default=None, alias="node_id"),
):
    start_dt, end_dt = parse_dt_param(start_dt), parse_dt_param(end_dt)
    base, params = build_base_sql(events, start_dt, end_dt, node_ids or [])
    if not base:
        return {"time_range": {"start": start_dt, "end": end_dt}, "events": []}

    sql = (
        f"SELECT event_type, COUNT(*) AS cnt "
        f"FROM ({base}) t "
        f"GROUP BY event_type ORDER BY cnt DESC"
    )
    rows = database.run_query(sql, params)
    return {
        "time_range": {"start": start_dt, "end": end_dt},
        "events": [{"event": r["event_type"], "count": r["cnt"]} for r in rows],
    }


@router.get("/list")
def get_list(
    start_dt: str,
    end_dt: str,
    events: List[str] = Query(default=ALL_EVENTS),
    node_ids: List[str] = Query(default=None, alias="node_id"),
):
    start_dt, end_dt = parse_dt_param(start_dt), parse_dt_param(end_dt)
    base, params = build_base_sql(events, start_dt, end_dt, node_ids or [])
    if not base:
        return {"records": []}

    sql = f"SELECT * FROM ({base}) t ORDER BY reg_dt DESC LIMIT 100"
    rows = database.run_query(sql, params)

    # node_id → name lookup from t_viewer_node
    name_map = {}
    if rows:
        ids = list({str(r["node_id"]) for r in rows})
        ph  = ",".join(["%s"] * len(ids))
        try:
            name_rows = database.run_query(
                f"SELECT DISTINCT ON (CAST(node_id AS TEXT)) "
                f"CAST(node_id AS TEXT) AS nid, name "
                f"FROM t_viewer_node WHERE CAST(node_id AS TEXT) IN ({ph})",
                ids,
            )
            name_map = {r["nid"]: r["name"] for r in name_rows}
        except Exception:
            pass

    # node_id → management_code lookup
    mgmt_map = {}
    if rows:
        ids = list({str(r["node_id"]) for r in rows})
        ph  = ",".join(["%s"] * len(ids))
        try:
            mc_rows = database.run_query(
                f"SELECT DISTINCT ON (CAST(node_id AS TEXT)) "
                f"CAST(node_id AS TEXT) AS nid, management_code "
                f"FROM t_viewer_node WHERE CAST(node_id AS TEXT) IN ({ph})",
                ids,
            )
            mgmt_map = {r["nid"]: r["management_code"] for r in mc_rows}
        except Exception:
            pass

    return {
        "records": [
            {
                "node_id":   r["node_id"],
                "node_name": name_map.get(str(r["node_id"]), ""),
                "ch":        r["ch"],
                "dtct_dt":   parse_dtct_dt(r.get("dtct_dt")),
                "event":     r["event_type"],
                "img_url":   img_to_api_url(r.get("img_path", "")),
                "thumb_url": img_to_thumb_url(r.get("img_path", "")),
                "img_path":  r.get("img_path", ""),
                "mgmt_code": mgmt_map.get(str(r["node_id"]), ""),
                "dst_val":   r.get("dst_val"),
            }
            for r in rows
        ]
    }


@router.get("/node-stats")
def get_node_stats(
    start_dt: str,
    end_dt: str,
    events: List[str] = Query(default=ALL_EVENTS),
    node_ids: List[str] = Query(default=None, alias="node_id"),
):
    start_dt, end_dt = parse_dt_param(start_dt), parse_dt_param(end_dt)
    base, params = build_base_sql(events, start_dt, end_dt, node_ids or [])
    if not base:
        return {"nodes": []}

    sql = (
        f"SELECT node_id, ch, COUNT(*) AS total "
        f"FROM ({base}) t "
        f"GROUP BY node_id, ch ORDER BY total DESC"
    )
    rows = database.run_query(sql, params)
    return {
        "nodes": [{"node_id": r["node_id"], "ch": r["ch"], "total": r["total"]} for r in rows]
    }


@router.get("/vlm-report")
def generate_vlm_report(
    node_id:    str,
    ch:         str,
    dtct_dt:    str,
    event_type: str,
    img_path:   str,
    cam_name:   str = "",
    mgmt_code:  str = "",
    dst_val:    float = None,
):
    """이벤트 이미지 + 메타데이터를 VLM 보고서 API로 전송하고 결과 반환."""
    import requests, base64, pathlib, mimetypes
    from typing import Optional

    win_path = linux_to_win(img_path)
    if not os.path.isfile(win_path):
        raise HTTPException(status_code=404, detail=f"Image not found: {win_path}")

    img_bytes = pathlib.Path(win_path).read_bytes()
    img_b64   = base64.b64encode(img_bytes).decode()
    filename  = pathlib.Path(win_path).name
    mime_type = mimetypes.guess_type(filename)[0] or "image/jpeg"

    event_json = {
        "event_type": event_type,
        "result":     "비정상",
        "event_time": f"UTC+0900:{dtct_dt}",
        "cam_name":   cam_name,
        "node_id":    int(node_id),
    }

    if event_type == "침수" and dst_val is not None:
        event_json["additional_info"] = {"침수율": dst_val}

    payload = {
        "event_json":       event_json,
        "image_base64":     img_b64,
        "image_filename":   filename,
        "image_mime_type":  mime_type,
    }

    try:
        resp = requests.post(REPORT_API_URL, json=payload, timeout=110)
        resp.raise_for_status()
        result = resp.json()
        result["_img_win_path"] = linux_to_win(img_path)
        result["_node_id"]     = node_id
        result["_event_type"]  = event_type
        return result
    except requests.exceptions.ConnectionError:
        raise HTTPException(status_code=502, detail="VLM API 서버에 연결할 수 없습니다.")
    except requests.exceptions.Timeout:
        raise HTTPException(status_code=504, detail="VLM API 응답 시간 초과.")
    except requests.exceptions.HTTPError as e:
        body = e.response.text[:500] if e.response is not None else ""
        raise HTTPException(status_code=502, detail=f"VLM API {e.response.status_code}: {body}")
    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=502, detail=str(e))


@router.post("/vlm-excel")
def generate_vlm_excel(result: dict = Body(...)):
    """VLM API 결과 JSON → Excel 파일 반환."""
    import openpyxl, io as _io
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    if not result.get("success"):
        raise HTTPException(status_code=400, detail="VLM 결과가 없습니다.")

    img_win_path = result.get("_img_win_path", "")
    rj       = result.get("report_json", {})
    ne       = result.get("normalized_event", {})
    obs      = result.get("observation_text", "")
    reporter = rj.get("보고자", {}) if isinstance(rj.get("보고자"), dict) else {}

    dt_str = rj.get("사고 발생일시", "")
    if "UTC+0900:" in dt_str:
        dt_str = dt_str.split("UTC+0900:")[-1].strip()

    rows = [
        ("보고자 성명",       reporter.get("성명", "")),
        ("보고자 근무조",     reporter.get("근무조", "")),
        ("카메라",            ne.get("cam_name", "")),
        ("Node ID",           ne.get("node_id", "")),
        ("이벤트 유형",       ne.get("event_type", "")),
        ("사고 발생일시",     dt_str),
        ("장소",              rj.get("장소", "")),
        ("관찰 내용",         obs),
        ("사고 관제내용",     rj.get("사고 관제내용", "")),
        ("피해 우려사항",     rj.get("피해 우려사항", "")),
        ("관제센터 조치사항", rj.get("관제센터 조치사항", "")),
        ("그 외 특이사항",    rj.get("그 외 특이사항", "")),
    ]

    def _row_h(val):
        s = str(val)
        weight = sum(2 if ord(c) > 0x2E7F else 1 for c in s)
        lines  = max(1, weight // 60 + s.count("\n") + 1)
        return max(18, min(lines * 18, 600))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "조치사항 보고서"

    thin   = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    wrap_l = Alignment(vertical="top", wrap_text=True)
    wrap_c = Alignment(horizontal="center", vertical="top", wrap_text=True)

    ws.merge_cells("A1:B1")
    hc = ws.cell(row=1, column=1, value="조치사항 보고서")
    hc.font      = Font(bold=True, size=14, color="FFFFFF")
    hc.fill      = PatternFill("solid", fgColor="1a4fa3")
    hc.alignment = center
    hc.border    = border
    ws.row_dimensions[1].height = 30

    for i, (label, value) in enumerate(rows):
        r  = i + 2
        lc = ws.cell(row=r, column=1, value=label)
        lc.font      = Font(bold=True)
        lc.fill      = PatternFill("solid", fgColor="D9E1F2")
        lc.alignment = wrap_l
        lc.border    = border
        vc = ws.cell(row=r, column=2, value=str(value))
        vc.alignment = wrap_l
        vc.border    = border
        ws.row_dimensions[r].height = _row_h(value)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 65

    # ── 이미지 썸네일 (하단) ─────────────────────────────────────────────────
    if img_win_path and os.path.isfile(img_win_path):
        try:
            from PIL import Image as _PIL
            from openpyxl.drawing.image import Image as _XLImg
            pil_img = _PIL.open(img_win_path)
            pil_img.thumbnail((320, 240), _PIL.LANCZOS)
            iw, ih = pil_img.size
            buf = _io.BytesIO()
            pil_img.save(buf, format="JPEG", quality=75)
            buf.seek(0)
            img_row = len(rows) + 2
            lc = ws.cell(row=img_row, column=1, value="이미지")
            lc.font      = Font(bold=True)
            lc.fill      = PatternFill("solid", fgColor="D9E1F2")
            lc.alignment = wrap_l
            lc.border    = border
            ws.cell(row=img_row, column=2).border = border
            ws.row_dimensions[img_row].height = int(ih * 0.75) + 6
            xl_img = _XLImg(buf)
            xl_img.width  = iw
            xl_img.height = ih
            ws.add_image(xl_img, f"B{img_row}")
        except Exception:
            pass

    event_nm = str(result.get("_event_type") or ne.get("event_type", "event")).strip() or "event"
    clean_dt = "".join(c if c.isdigit() else "_" for c in dt_str).strip("_") or datetime.now().strftime("%Y%m%d_%H%M%S")
    fname    = f"{event_nm}_{clean_dt}.xlsx"
    path     = os.path.join(tempfile.gettempdir(), fname)
    wb.save(path)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=fname,
    )


@router.get("/node-detail")
def get_node_detail(
    node_id: str,
    ch: str,
    start_dt: str,
    end_dt: str,
    events: List[str] = Query(default=ALL_EVENTS),
):
    start_dt, end_dt = parse_dt_param(start_dt), parse_dt_param(end_dt)
    base, params = build_base_sql(events, start_dt, end_dt)
    if not base:
        return {"node_id": node_id, "ch": ch, "events": []}

    sql = (
        f"SELECT event_type, COUNT(*) AS cnt "
        f"FROM ({base}) t "
        f"WHERE CAST(node_id AS TEXT) = %s AND CAST(ch AS TEXT) = %s "
        f"GROUP BY event_type ORDER BY cnt DESC"
    )
    rows = database.run_query(sql, params + [node_id, ch])
    return {
        "node_id": node_id,
        "ch": ch,
        "events": [{"event": r["event_type"], "count": r["cnt"]} for r in rows],
    }
