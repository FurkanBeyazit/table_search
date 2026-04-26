import matplotlib.pyplot as plt
import requests
import gradio as gr
from config import ALL_EVENTS, BHVR_EVENTS, DST_EVENTS, API_BASE_URL
from ui_charts import (build_histogram, build_line_chart, build_server_line,
    build_server_histogram, build_precision_bar, build_precision_trend,
    build_precision_count_trend, build_false_cause_event_chart, build_time_heatmap,
    build_time_slot_bar, build_time_line, build_period_chart,
    build_operator_chart_trend)
from ui_render import (render_today_events, render_summary_counts, render_detail_table,
    render_nodes_table, render_server_stats, render_precision_cards,
    render_precision_event_table, render_precision_node_table, render_stats, render_list,
    render_node_stats, render_false_cause_completion, render_false_cause_event_table,
    render_false_cause_user_table, render_time_dist_cards, render_period_list,
    render_operator_table, render_operator_daily_table, render_operator_30day_table,
    render_operator_monthly_table)


def api_get(path: str, params: dict = None) -> dict:
    try:
        r = requests.get(f"{API_BASE_URL}{path}", params=params or {}, timeout=30)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        return {"error": str(e)}


def load_today_tab():
    today_html = render_today_events(api_get("/api/stats/today"))
    histogram  = build_histogram(api_get("/api/stats/histogram"))
    return today_html, histogram


def load_summary_tab():
    data      = api_get("/api/stats/summary")
    empty_fig = plt.figure(figsize=(20, 4))

    if "error" in data:
        err = f"<p style='color:red'>⚠ {data['error']}</p>"
        return err, empty_fig, err, err, empty_fig, err

    bhvr = data.get("bhvr", {})
    dst  = data.get("dst",  {})

    return (
        render_summary_counts(bhvr.get("summary", {}), "🚶 행동 분석 (bhvr) 요약"),
        build_line_chart(bhvr.get("line", []), "행동 분석 일별 추이 (30일)"),
        render_detail_table(bhvr.get("detail", []), BHVR_EVENTS),
        render_summary_counts(dst.get("summary",  {}), "🌊 재난 분석 (dst) 요약"),
        build_line_chart(dst.get("line",  []), "재난 분석 일별 추이 (30일)"),
        render_detail_table(dst.get("detail",  []), DST_EVENTS),
    )


def get_viewer_names() -> list:
    data = api_get("/api/server/nodes")
    nodes = data.get("nodes", [])
    return sorted(set(n["viewer_name"] for n in nodes if n.get("viewer_name")))


def do_load_nodes_by_viewer(viewer_name: str) -> str:
    data  = api_get("/api/server/nodes")
    nodes = data.get("nodes", [])
    if viewer_name:
        nodes = [n for n in nodes if n.get("viewer_name") == viewer_name]
    return render_nodes_table(nodes)


def do_import_excel(file_obj):
    if file_obj is None:
        return "<p style='color:orange'>파일을 선택하세요</p>", gr.update(), ""
    file_path = file_obj.name if hasattr(file_obj, "name") else file_obj
    try:
        with open(file_path, "rb") as f:
            r = requests.post(
                f"{API_BASE_URL}/api/server/import",
                files={"file": ("upload.xlsx", f,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=30,
            )
        if r.ok:
            n       = r.json().get("imported", 0)
            viewers = get_viewer_names()
            first   = viewers[0] if viewers else None
            nodes_html = do_load_nodes_by_viewer(first) if first else ""
            return (
                f"<p style='color:green'>✓ {n}건 임포트 완료 &nbsp;·&nbsp; 뷰어 {len(viewers)}개</p>",
                gr.update(choices=viewers, value=first),
                nodes_html,
            )
        return f"<p style='color:red'>⚠ {r.text}</p>", gr.update(), ""
    except Exception as e:
        return f"<p style='color:red'>⚠ {e}</p>", gr.update(), ""


def do_add_node(viewer, node_id, mgmt, name):
    if not viewer.strip() or not node_id.strip():
        return "<p style='color:orange'>Viewer Name과 Node ID는 필수입니다</p>", do_load_nodes_by_viewer(viewer)
    try:
        r = requests.post(
            f"{API_BASE_URL}/api/server/nodes",
            json={"viewer_name": viewer, "node_id": node_id,
                  "management_code": mgmt, "name": name},
            timeout=10,
        )
        if r.ok:
            return "<p style='color:green'>✓ 추가 완료</p>", do_load_nodes_by_viewer(viewer)
        return f"<p style='color:red'>⚠ {r.text}</p>", do_load_nodes_by_viewer(viewer)
    except Exception as e:
        return f"<p style='color:red'>⚠ {e}</p>", do_load_nodes_by_viewer(viewer)


def do_load_server_stats():
    data = api_get("/api/server/stats")
    return render_server_stats(data), build_server_line(data), build_server_histogram(data)


def do_load_precision(period: str):
    data      = api_get("/api/analysis/precision", {"period": period})
    empty_fig = plt.figure(figsize=(20, 4))
    if "error" in data:
        err = f"<p style='color:red'>⚠ {data['error']}</p>"
        return err, empty_fig, err, err, empty_fig, empty_fig
    daily = data.get("daily", [])
    return (
        render_precision_cards(data.get("summary", {})),
        build_precision_bar(data.get("events", [])),
        render_precision_event_table(data.get("events", [])),
        render_precision_node_table(data.get("nodes", [])),
        build_precision_trend(daily),
        build_precision_count_trend(daily),
    )


def do_search(start_dt: str, end_dt: str, selected_events: list, node_id_input: str):
    if not selected_events:
        msg = "<p style='color:orange'>⚠ 이벤트를 하나 이상 선택하세요</p>"
        return msg, msg, msg

    node_ids = [n.strip() for n in node_id_input.split(",") if n.strip()] if node_id_input else []

    params = {"start_dt": start_dt, "end_dt": end_dt, "events": selected_events}
    if node_ids:
        params["node_id"] = node_ids

    stats_html = render_stats(api_get("/api/search/stats", params))
    list_html  = render_list(api_get("/api/search/list",   params))
    node_html  = render_node_stats(
        api_get("/api/search/node-stats", params),
        start_dt, end_dt, selected_events,
    )
    return stats_html, list_html, node_html


def do_load_false_cause(period: str):
    try:
        data      = api_get("/api/analysis/false_cause", {"period": period})
        empty_fig = plt.figure(figsize=(10, 6))
        if "error" in data:
            err = f"<p style='color:red'>⚠ {data['error']}</p>"
            return err, empty_fig, err, err
        all_causes = data.get("all_causes", [])
        return (
            render_false_cause_completion(data.get("completion", {})),
            build_false_cause_event_chart(data.get("events", []), all_causes),
            render_false_cause_event_table(data.get("events", []), all_causes),
            render_false_cause_user_table(data.get("users",  []), all_causes),
        )
    except Exception as e:
        import traceback
        err = f"<p style='color:red'>⚠ <pre>{traceback.format_exc()}</pre></p>"
        return err, plt.figure(figsize=(10, 6)), err, err


def do_load_time_dist_all(period: str):
    try:
        data      = api_get("/api/analysis/time_dist_all", {"period": period})
        empty_fig = plt.figure(figsize=(20, 4))
        if "error" in data:
            err = f"<p style='color:red'>⚠ {data['error']}</p>"
            return err, empty_fig, empty_fig, empty_fig
        return (
            render_time_dist_cards(data.get("cards", {}), total_label="이벤트 총계"),
            build_time_heatmap(data.get("hourly_events", {}), title="시간대 × 이벤트 히트맵 (전체)"),
            build_time_line(data.get("hour_total", []), title="시간별 이벤트 분포 (0~23시)"),
            build_time_slot_bar(data.get("slots", [])),
        )
    except Exception as e:
        import traceback
        err = f"<p style='color:red'>⚠ <pre>{traceback.format_exc()}</pre></p>"
        return err, plt.figure(figsize=(20, 4)), plt.figure(figsize=(20, 4)), plt.figure(figsize=(20, 4))


def do_load_time_dist(period: str):
    try:
        data      = api_get("/api/analysis/time_dist", {"period": period})
        empty_fig = plt.figure(figsize=(20, 4))
        if "error" in data:
            err = f"<p style='color:red'>⚠ {data['error']}</p>"
            return err, empty_fig, empty_fig, empty_fig
        return (
            render_time_dist_cards(data.get("cards", {})),
            build_time_heatmap(data.get("hourly_events", {})),
            build_time_line(data.get("hour_total", [])),
            build_time_slot_bar(data.get("slots", [])),
        )
    except Exception as e:
        import traceback
        err = f"<p style='color:red'>⚠ <pre>{traceback.format_exc()}</pre></p>"
        return err, plt.figure(figsize=(20, 4)), plt.figure(figsize=(20, 4)), plt.figure(figsize=(20, 4))


def do_period_query(ref_date: str, time_from: str, time_to: str, event: str):
    params = {
        "ref_date":  (ref_date or "").strip(),
        "time_from": (time_from or "00:00").strip(),
        "time_to":   (time_to   or "23:59").strip(),
        "event":     event or "전체",
    }
    data      = api_get("/api/stats/period_query", params)
    empty_fig = plt.figure(figsize=(14, 5))
    if "error" in data:
        err = f"<p style='color:red'>⚠ {data['error']}</p>"
        return empty_fig, err
    return build_period_chart(data), render_period_list(data)


def do_export_list(start_dt: str, end_dt: str, selected_events: list, node_id_input: str):
    """조희 List 탭 결과를 xlsx로 내보냄. 이미지 URL은 하이퍼링크로."""
    import openpyxl, tempfile
    from openpyxl.styles import Font

    node_ids = [n.strip() for n in node_id_input.split(",") if n.strip()] if node_id_input else []
    params   = {"start_dt": start_dt, "end_dt": end_dt, "events": selected_events}
    if node_ids:
        params["node_id"] = node_ids

    data    = api_get("/api/search/list", params)
    records = data.get("records", [])
    if not records:
        return gr.update()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "조희결과"
    ws.append(["Node ID", "Name", "Ch", "Detect Time", "Event", "이미지"])

    for rec in records:
        row_idx = ws.max_row + 1
        ws.append([
            rec.get("node_id", ""),
            rec.get("node_name", ""),
            rec.get("ch", ""),
            rec.get("dtct_dt", ""),
            rec.get("event", ""),
            "",
        ])
        img_url = rec.get("img_url", "")
        if img_url:
            cell = ws.cell(row=row_idx, column=6, value="이미지 보기")
            cell.hyperlink = img_url
            cell.font = Font(color="0563C1", underline="single")

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix="list_export_")
    wb.save(tmp.name)
    tmp.close()
    return gr.update(visible=True, value=tmp.name)


def _export_summary_excel(table_type: str):
    """bhvr veya dst detail tablosunu xlsx olarak oluşturur."""
    import openpyxl, tempfile
    data = api_get("/api/stats/summary")
    if "error" in data:
        return gr.update()
    detail      = data.get(table_type, {}).get("detail", [])
    events_list = BHVR_EVENTS if table_type == "bhvr" else DST_EVENTS
    if not detail:
        return gr.update()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "30일 상세"
    ws.append(["날짜", "합계"] + events_list)
    for row in detail:
        ws.append(
            [f"{row['date'][5:]}({row['label']})", row.get("total", 0)]
            + [row.get(ev, 0) for ev in events_list]
        )
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False, prefix=f"{table_type}_export_")
    wb.save(tmp.name)
    tmp.close()
    return gr.update(visible=True, value=tmp.name)


def do_export_bhvr():
    return _export_summary_excel("bhvr")


def do_export_dst():
    return _export_summary_excel("dst")


def do_load_operator_init():
    """앱 시작 시 두 탭 모두 초기화: dropdown1 + chart + daily_table + dropdown2 + detail_table."""
    empty_fig  = plt.figure(figsize=(20, 9))
    empty_html = "<p style='opacity:0.5'>운영자를 선택하세요</p>"
    data = api_get("/api/analysis/operator_summary")
    if "error" in data:
        err = f"<p style='color:red'>⚠ {data['error']}</p>"
        return (
            gr.update(choices=[]),
            empty_fig, empty_html,
            gr.update(choices=["전체 보기"]),
            err,
        )
    operators      = [op["reg_id"] for op in data.get("operators", [])]
    first          = operators[0] if operators else None
    chart_fig, daily_table, monthly_table = (
        do_load_operator_chart(first) if first else (empty_fig, empty_html, empty_html)
    )
    detail_choices = ["전체 보기"] + operators
    return (
        gr.update(choices=operators, value=first),
        chart_fig,
        daily_table,
        monthly_table,
        gr.update(choices=detail_choices, value="전체 보기"),
        do_load_operator_detail("전체 보기"),
    )


def do_load_operator_chart(reg_id: str):
    """14일 그래프 + 14일 일별 테이블 + 연간 월별 테이블 반환."""
    empty_fig  = plt.figure(figsize=(20, 9))
    empty_html = "<p style='opacity:0.5'>운영자를 선택하세요</p>"
    if not reg_id:
        return empty_fig, empty_html, empty_html
    try:
        data = api_get("/api/analysis/operator_chart", {"reg_id": reg_id})
        if "error" in data:
            return empty_fig, empty_html, empty_html
        return (
            build_operator_chart_trend(data),
            render_operator_daily_table(data),
            render_operator_monthly_table(data),
        )
    except Exception:
        return empty_fig, empty_html, empty_html


def do_load_operator_detail(reg_id: str):
    """비교 테이블 항상 표시. 특정 운영자 선택 시 highlight + 30일 일별 테이블 추가."""
    summary_data = api_get("/api/analysis/operator_summary")
    if "error" in summary_data:
        return f"<p style='color:red'>⚠ {summary_data['error']}</p>"

    highlight = reg_id if reg_id and reg_id != "전체 보기" else None
    html = render_operator_table(summary_data, highlight_reg_id=highlight)

    if highlight:
        chart_data = api_get("/api/analysis/operator_chart", {"reg_id": reg_id})
        if "error" not in chart_data:
            html += "<h4 style='margin:20px 0 8px'>📅 30일 일별 현황</h4>"
            html += render_operator_30day_table(chart_data)

    return html


def do_generate_monthly_report(year: int, month: int):
    """월간 보고서 Excel 파일 생성."""
    import openpyxl, tempfile, calendar
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = api_get("/api/analysis/monthly_report", {"year": year, "month": month})
    if "error" in data:
        return gr.update(), f"<p style='color:red'>⚠ {data['error']}</p>"

    days       = data["days"]          # ["2026-03-01", ...]
    all_events = data["all_events"]
    ev_day     = data["ev_day"]        # {event: {day_str: {jeongdam, odam}}}
    cameras    = data["cameras"]       # [{node_id, ch, name}, ...]
    cam_day    = data["cam_day"]       # {"node_ch": {day_str: {jeongdam, odam}}}
    ev_cam_day = data["ev_cam_day"]    # {event: {"node_ch": {day_str: {jeongdam, odam}}}}

    n_days = len(days)
    day_labels = [d[5:].replace("-", "/") for d in days]  # "03/01"

    # ─── 공통 스타일 ─────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
    HEADER2_FILL = PatternFill("solid", fgColor="8EA9DB")
    TOTAL_FILL   = PatternFill("solid", fgColor="D9E1F2")
    SUMROW_FILL  = PatternFill("solid", fgColor="FCE4D6")
    WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
    GRAY_FILL    = PatternFill("solid", fgColor="F2F2F2")

    header_font  = Font(bold=True, color="FFFFFF")
    header2_font = Font(bold=True, color="000000")
    bold         = Font(bold=True)
    thin = Side(border_style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center")

    def _hdr(ws, row, col, val, fill=None, font=None, align=None):
        c = ws.cell(row=row, column=col, value=val)
        if fill:  c.fill = fill
        if font:  c.font = font
        c.alignment = align or center
        c.border    = border
        return c

    def _val(ws, row, col, val, fill=None, font=None, align=None):
        c = ws.cell(row=row, column=col, value=val)
        c.border    = border
        c.alignment = align or center
        if fill: c.fill = fill
        if font: c.font = font
        return c

    def _merge(ws, r1, c1, r2, c2, val, fill=None, font=None, align=None):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cell = ws.cell(row=r1, column=c1, value=val)
        cell.alignment = align or center
        cell.border    = border
        if fill: cell.fill = fill
        if font: cell.font = font
        for r in range(r1, r2+1):
            for c_ in range(c1, c2+1):
                ws.cell(row=r, column=c_).border = border

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET 1: 전체 (event × day)
    # 열 구조: [이벤트 | 03/01 정탐 | 03/01 오탐 | 03/02 정탐 | ... | 월합계 정탐 | 월합계 오탐 | 월합계]
    # ─────────────────────────────────────────────────────────────────────────
    def build_sheet1(wb):
        ws = wb.active
        ws.title = "전체"

        LEFT_COLS = 1   # 이벤트

        # Row 1: 날짜 머리글 (병합)
        _merge(ws, 1, 1, 2, 1, "이벤트", fill=HEADER_FILL, font=header_font)
        for i, dlabel in enumerate(day_labels):
            col = LEFT_COLS + 1 + i * 2
            _merge(ws, 1, col, 1, col+1, dlabel, fill=HEADER_FILL, font=header_font)
        total_col_start = LEFT_COLS + 1 + n_days * 2
        _merge(ws, 1, total_col_start, 1, total_col_start+1, "월합계", fill=HEADER_FILL, font=header_font)

        # Row 2: 정탐/오탐 반복
        for i in range(n_days):
            col = LEFT_COLS + 1 + i * 2
            _hdr(ws, 2, col,   "정탐", fill=HEADER2_FILL, font=header2_font)
            _hdr(ws, 2, col+1, "오탐", fill=HEADER2_FILL, font=header2_font)
        _hdr(ws, 2, total_col_start,   "정탐", fill=TOTAL_FILL, font=bold)
        _hdr(ws, 2, total_col_start+1, "오탐", fill=TOTAL_FILL, font=bold)

        # Data rows
        month_jd_by_day = [0] * n_days
        month_od_by_day = [0] * n_days

        for row_idx, ev in enumerate(all_events):
            r = 3 + row_idx
            fill = GRAY_FILL if row_idx % 2 == 1 else WHITE_FILL
            _val(ws, r, 1, ev, fill=fill, align=left)
            ev_data = ev_day.get(ev, {})
            total_jd = total_od = 0
            for i, ds in enumerate(days):
                col = LEFT_COLS + 1 + i * 2
                jd = ev_data.get(ds, {}).get("jeongdam", 0)
                od = ev_data.get(ds, {}).get("odam", 0)
                _val(ws, r, col,   jd or 0, fill=fill)
                _val(ws, r, col+1, od or 0, fill=fill)
                total_jd += jd
                total_od += od
                month_jd_by_day[i] += jd
                month_od_by_day[i] += od
            _val(ws, r, total_col_start,   total_jd or 0, fill=fill, font=bold)
            _val(ws, r, total_col_start+1, total_od or 0, fill=fill, font=bold)

        # Summary rows at bottom
        sum_row = 3 + len(all_events)
        _val(ws, sum_row, 1, "정탐합계", fill=SUMROW_FILL, font=bold, align=left)
        grand_jd = 0
        for i in range(n_days):
            col = LEFT_COLS + 1 + i * 2
            _val(ws, sum_row, col,   month_jd_by_day[i], fill=SUMROW_FILL, font=bold)
            _val(ws, sum_row, col+1, "",                 fill=SUMROW_FILL)
            grand_jd += month_jd_by_day[i]
        _val(ws, sum_row, total_col_start,   grand_jd, fill=SUMROW_FILL, font=bold)
        _val(ws, sum_row, total_col_start+1, "",        fill=SUMROW_FILL)

        od_row = sum_row + 1
        _val(ws, od_row, 1, "오탐합계", fill=SUMROW_FILL, font=bold, align=left)
        grand_od = 0
        for i in range(n_days):
            col = LEFT_COLS + 1 + i * 2
            _val(ws, od_row, col,   "",                 fill=SUMROW_FILL)
            _val(ws, od_row, col+1, month_od_by_day[i], fill=SUMROW_FILL, font=bold)
            grand_od += month_od_by_day[i]
        _val(ws, od_row, total_col_start,   "",       fill=SUMROW_FILL)
        _val(ws, od_row, total_col_start+1, grand_od, fill=SUMROW_FILL, font=bold)

        total_row = od_row + 1
        _val(ws, total_row, 1, "전체합계", fill=SUMROW_FILL, font=bold, align=left)
        for i in range(n_days):
            col = LEFT_COLS + 1 + i * 2
            day_total = month_jd_by_day[i] + month_od_by_day[i]
            _merge(ws, total_row, col, total_row, col+1, day_total, fill=SUMROW_FILL, font=bold)
        grand_total = grand_jd + grand_od
        _merge(ws, total_row, total_col_start, total_row, total_col_start+1, grand_total, fill=SUMROW_FILL, font=bold)

        # column widths
        ws.column_dimensions["A"].width = 14
        for i in range(n_days * 2 + 2):
            ws.column_dimensions[get_column_letter(2 + i)].width = 7
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

    # ─────────────────────────────────────────────────────────────────────────
    # SHEET builder: camera rows (used for 카메라 sheet and per-event sheets)
    # LEFT_COLS = 3: [Node ID | Ch | 카메라명]
    # ─────────────────────────────────────────────────────────────────────────
    def build_camera_sheet(wb, sheet_title, day_data_map):
        """
        day_data_map: {cam_key("nodeid_ch"): {day_str: {jeongdam, odam}}}
        """
        ws = wb.create_sheet(title=sheet_title)
        LEFT_COLS = 3

        _merge(ws, 1, 1, 2, 1, "Node ID", fill=HEADER_FILL, font=header_font)
        _merge(ws, 1, 2, 2, 2, "Ch",      fill=HEADER_FILL, font=header_font)
        _merge(ws, 1, 3, 2, 3, "카메라명",  fill=HEADER_FILL, font=header_font)
        for i, dlabel in enumerate(day_labels):
            col = LEFT_COLS + 1 + i * 2
            _merge(ws, 1, col, 1, col+1, dlabel, fill=HEADER_FILL, font=header_font)
        total_col_start = LEFT_COLS + 1 + n_days * 2
        _merge(ws, 1, total_col_start, 1, total_col_start+1, "월합계", fill=HEADER_FILL, font=header_font)

        for i in range(n_days):
            col = LEFT_COLS + 1 + i * 2
            _hdr(ws, 2, col,   "정탐", fill=HEADER2_FILL, font=header2_font)
            _hdr(ws, 2, col+1, "오탐", fill=HEADER2_FILL, font=header2_font)
        _hdr(ws, 2, total_col_start,   "정탐", fill=TOTAL_FILL, font=bold)
        _hdr(ws, 2, total_col_start+1, "오탐", fill=TOTAL_FILL, font=bold)

        month_jd_by_day = [0] * n_days
        month_od_by_day = [0] * n_days

        for row_idx, cam in enumerate(cameras):
            r    = 3 + row_idx
            fill = GRAY_FILL if row_idx % 2 == 1 else WHITE_FILL
            cam_key = f"{cam['node_id']}_{cam['ch']}"
            _val(ws, r, 1, cam["node_id"], fill=fill, align=left)
            _val(ws, r, 2, cam["ch"],      fill=fill)
            _val(ws, r, 3, cam["name"],    fill=fill, align=left)
            cd = day_data_map.get(cam_key, {})
            total_jd = total_od = 0
            for i, ds in enumerate(days):
                col = LEFT_COLS + 1 + i * 2
                jd = cd.get(ds, {}).get("jeongdam", 0)
                od = cd.get(ds, {}).get("odam", 0)
                _val(ws, r, col,   jd or 0, fill=fill)
                _val(ws, r, col+1, od or 0, fill=fill)
                total_jd += jd
                total_od += od
                month_jd_by_day[i] += jd
                month_od_by_day[i] += od
            _val(ws, r, total_col_start,   total_jd or 0, fill=fill, font=bold)
            _val(ws, r, total_col_start+1, total_od or 0, fill=fill, font=bold)

        # summary rows
        sum_row = 3 + len(cameras)
        _val(ws, sum_row, 1, "정탐합계", fill=SUMROW_FILL, font=bold)
        ws.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=3)
        grand_jd = 0
        for i in range(n_days):
            col = LEFT_COLS + 1 + i * 2
            _val(ws, sum_row, col,   month_jd_by_day[i], fill=SUMROW_FILL, font=bold)
            _val(ws, sum_row, col+1, "",                 fill=SUMROW_FILL)
            grand_jd += month_jd_by_day[i]
        _val(ws, sum_row, total_col_start,   grand_jd, fill=SUMROW_FILL, font=bold)
        _val(ws, sum_row, total_col_start+1, "",        fill=SUMROW_FILL)

        od_row = sum_row + 1
        _val(ws, od_row, 1, "오탐합계", fill=SUMROW_FILL, font=bold)
        ws.merge_cells(start_row=od_row, start_column=1, end_row=od_row, end_column=3)
        grand_od = 0
        for i in range(n_days):
            col = LEFT_COLS + 1 + i * 2
            _val(ws, od_row, col,   "",                 fill=SUMROW_FILL)
            _val(ws, od_row, col+1, month_od_by_day[i], fill=SUMROW_FILL, font=bold)
            grand_od += month_od_by_day[i]
        _val(ws, od_row, total_col_start,   "",       fill=SUMROW_FILL)
        _val(ws, od_row, total_col_start+1, grand_od, fill=SUMROW_FILL, font=bold)

        total_row = od_row + 1
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        _val(ws, total_row, 1, "전체합계", fill=SUMROW_FILL, font=bold)
        for i in range(n_days):
            col = LEFT_COLS + 1 + i * 2
            day_total = month_jd_by_day[i] + month_od_by_day[i]
            _merge(ws, total_row, col, total_row, col+1, day_total, fill=SUMROW_FILL, font=bold)
        _merge(ws, total_row, total_col_start, total_row, total_col_start+1,
               grand_jd + grand_od, fill=SUMROW_FILL, font=bold)

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 6
        ws.column_dimensions["C"].width = 18
        for i in range(n_days * 2 + 2):
            ws.column_dimensions[get_column_letter(4 + i)].width = 7
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18

    # ─────────────────────────────────────────────────────────────────────────
    # BUILD WORKBOOK
    # ─────────────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    build_sheet1(wb)
    build_camera_sheet(wb, "카메라", cam_day)
    for ev in all_events:
        sheet_ev = ev_cam_day.get(ev, {})
        build_camera_sheet(wb, ev, sheet_ev)

    tmp = tempfile.NamedTemporaryFile(
        suffix=".xlsx", delete=False,
        prefix=f"monthly_{year}{month:02d}_"
    )
    wb.save(tmp.name)
    tmp.close()
    return gr.update(visible=True, value=tmp.name), "<p style='color:green'>✓ 보고서 생성 완료</p>"
